import openpyxl
import pandas as pd 

wb = openpyxl.load_workbook(r'C:\Users\Nisserine\Desktop\GOAL\Day 11\13\practice_quotations.xlsx', data_only=True)

seen_quotations = {}  
duplicates = []       
failed_sheets = []    

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    try:
        customer_name = ws.cell(row=9, column=4).value
        quotation_date = ws.cell(row=10, column=4).value
        
        if quotation_date is None or customer_name is None:
            failed_sheets.append({
                'sheet_name': sheet_name,
                'issue': 'ورقة فارغة أو غير مطابقة للقالب (Missing Date/Name)'
            })
            continue
            
        item_count = 0
        for r in range(14, ws.max_row + 1):
            if ws.cell(row=r, column=2).value is not None:
                item_count += 1
                
        signature = (customer_name, str(quotation_date), item_count)
        
        if signature in seen_quotations:
            original_sheet = seen_quotations[signature]
            duplicates.append({
                'duplicate_sheet': sheet_name,
                'original_sheet': original_sheet,
                'customer': customer_name,
                'status': 'مكرر - يتطلب التجاهل عند التجميع'
            })
        else:
            seen_quotations[signature] = sheet_name

    except Exception as e:
        failed_sheets.append({
            'sheet_name': sheet_name,
            'issue': f'خطأ أثناء القراءة: {str(e)}'
        })

df_duplicates = pd.DataFrame(duplicates)
df_failed = pd.DataFrame(failed_sheets)

with pd.ExcelWriter(r'C:\Users\Nisserine\Desktop\GOAL\Day 11\13\audit_report.xlsx', engine='openpyxl') as writer:
    if not df_duplicates.empty:
        df_duplicates.to_excel(writer, sheet_name='المكررات (Duplicates)', index=False)
    if not df_failed.empty:
        df_failed.to_excel(writer, sheet_name='الأوراق الشاذة (Failed)', index=False)


